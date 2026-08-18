#!/usr/bin/env python3
"""
convert.py  —  Excel Hiring Sheet  ->  data.json
=================================================

Converts the "New Joiners Cost" Excel workbook into the JSON structure consumed
by the HR Hiring Cost & Vacancy Management System (index.html).

The Excel sheet is the source of truth for the field structure. Cost columns are
bucketed by their SPREADSHEET COLUMN LETTER (not by header text), exactly matching
the cost-frequency rules the business uses:

    Columns M-O  -> one-time costs        (charged once, in the hiring month)
    Columns P-R  -> monthly recurring     (every month from the hiring month)
    Column  S    -> once every 2 years    (hiring month, then every 24 months)
    Columns T-X  -> annual costs          (hiring month, then every 12 months)
    Column  Y    -> monthly cost          (every month from the hiring month)
    Column  Z    -> two-year cost (Gratuity); treated as a 2-year bucket item

The final "2-Year Employment Cost" (sheet column AB "Total") is NOT trusted as a
manual value — the application recalculates it from the rules above.

Usage
-----
    python convert.py hiring.xlsx
    python convert.py hiring.xlsx custom-data.json      # optional output name

Requires: pandas, openpyxl  (pip install pandas openpyxl)
"""

import sys
import json
import math
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl is required.  Run:  pip install openpyxl pandas")


# ---------------------------------------------------------------------------
# Column layout (1-based spreadsheet column indexes).
# These map the physical columns to their meaning. If the sheet layout changes,
# adjust here — everything else derives from this map.
# ---------------------------------------------------------------------------
COL = {
    "srNo": 1,        # A  Sr No.
    "employee": 2,    # B  Employee Name
    "qty": 3,         # C  QTY
    "position": 4,    # D  Job Title
    "company": 5,     # E  Company
    "department": 6,  # F  Department
    "collar": 7,      # G  Type (Blue/White collar)
    "employment": 8,  # H  Employment Type (Permanent/Contractual)
    "date": 9,        # I  DATE  (hiring / start date)
    "month": 10,      # J  MONTH
    "localOverseas": 11,  # K
    "source": 12,     # L
}

# Cost buckets by spreadsheet column letter -> (start_col, end_col) inclusive.
ONE_TIME_COLS   = (13, 15)   # M, N, O
MONTHLY_REC_COLS = (16, 18)  # P, Q, R
TWO_YEAR_COL    = 19         # S
ANNUAL_COLS     = (20, 24)   # T, U, V, W, X
MONTHLY_COL     = 25         # Y
GRATUITY_COL    = 26         # Z  (two-year bucket)
YRS_COL         = 27         # AA
TOTAL_COL       = 28         # AB (recalculated by the app; kept for reference)
REMARKS_COL     = 29         # AC


def clean_header(text):
    """Collapse the multi-line Excel headers into a single clean label."""
    if text is None:
        return ""
    return " ".join(str(text).replace("\n", " ").split()).strip()


def to_number(value):
    """Convert a cell to a float, treating blanks / NaN / text as 0.0."""
    if value is None:
        return 0.0
    if isinstance(value, str):
        v = value.strip().replace(",", "")
        if v == "":
            return 0.0
        try:
            return float(v)
        except ValueError:
            return 0.0
    if isinstance(value, float) and math.isnan(value):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def to_iso_date(value):
    """Return an ISO date string (YYYY-MM-DD) or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def bucket(ws, header_map, row, start, end):
    """Build {header: amount} for an inclusive column range."""
    out = {}
    for c in range(start, end + 1):
        label = header_map[c]
        out[label] = to_number(cell(ws, row, c))
    return out


def convert(input_path, output_path):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row

    # ---- Header map: column index -> clean label -------------------------
    header_map = {}
    columns_meta = []
    for c in range(1, max_col + 1):
        label = clean_header(cell(ws, 1, c))
        header_map[c] = label
        columns_meta.append({"letter": get_column_letter(c), "index": c, "label": label})

    # ---- Validate required columns ---------------------------------------
    required = {
        "Company": COL["company"],
        "Department": COL["department"],
        "Job Title": COL["position"],
        "DATE": COL["date"],
    }
    missing = [name for name, idx in required.items()
               if clean_header(cell(ws, 1, idx)).lower() != name.lower()
               and name.lower() not in clean_header(cell(ws, 1, idx)).lower()]
    if missing:
        print(f"WARNING: expected columns not found where anticipated: {missing}")
        print("         Proceeding with positional mapping — verify the output.")

    warnings = []
    vacancies = []
    seen_keys = {}
    year_counter = {}

    for row in range(2, max_row + 1):
        # Skip completely empty rows.
        if all(cell(ws, row, c) in (None, "") for c in range(1, max_col + 1)):
            continue

        company = (cell(ws, row, COL["company"]) or "").strip() if isinstance(cell(ws, row, COL["company"]), str) else cell(ws, row, COL["company"])
        if isinstance(company, str):
            company = COMPANY_RENAMES.get(company, company)
        department = cell(ws, row, COL["department"])
        position = cell(ws, row, COL["position"])
        hiring_iso = to_iso_date(cell(ws, row, COL["date"]))
        employee = cell(ws, row, COL["employee"])
        qty_raw = cell(ws, row, COL["qty"])
        qty = int(to_number(qty_raw)) if to_number(qty_raw) > 0 else 1

        # ---- Data-quality checks (report, don't drop) --------------------
        rlabel = f"row {row}"
        if not company:
            warnings.append(f"{rlabel}: missing Company")
        if not department:
            warnings.append(f"{rlabel}: missing Department")
        if not position:
            warnings.append(f"{rlabel}: missing Job Title")
        if not hiring_iso:
            warnings.append(f"{rlabel}: missing/invalid hiring DATE")

        # ---- Unique vacancy ID:  VAC-<year>-<seq> ------------------------
        year = hiring_iso[:4] if hiring_iso else datetime.now().strftime("%Y")
        year_counter[year] = year_counter.get(year, 0) + 1
        vac_id = f"VAC-{year}-{year_counter[year]:03d}"

        # ---- Duplicate detection (same company/pos/dept/date) ------------
        dkey = (str(company), str(department), str(position), str(hiring_iso))
        if dkey in seen_keys:
            warnings.append(f"{rlabel}: possible duplicate of {seen_keys[dkey]} "
                            f"({position} / {company} / {hiring_iso})")
        else:
            seen_keys[dkey] = vac_id

        # ---- Status: sheet has no status column. Blank Employee Name => a
        #      planned/open vacancy; a filled name => treat as Hired. -------
        has_name = employee not in (None, "") and not (isinstance(employee, float) and math.isnan(employee))
        status = "Hired" if has_name else "Planned"

        # ---- Cost buckets by column letter -------------------------------
        one_time = bucket(ws, header_map, row, *ONE_TIME_COLS)
        monthly_rec = bucket(ws, header_map, row, *MONTHLY_REC_COLS)
        annual = bucket(ws, header_map, row, *ANNUAL_COLS)
        two_year = {
            header_map[TWO_YEAR_COL]: to_number(cell(ws, row, TWO_YEAR_COL)),
            header_map[GRATUITY_COL]: to_number(cell(ws, row, GRATUITY_COL)),
        }
        monthly = {header_map[MONTHLY_COL]: to_number(cell(ws, row, MONTHLY_COL))}

        # ---- Original row values (preserve everything) -------------------
        original = {}
        for c in range(1, max_col + 1):
            v = cell(ws, row, c)
            if isinstance(v, (datetime, date)):
                v = v.isoformat()
            elif isinstance(v, float) and math.isnan(v):
                v = None
            original[header_map[c]] = v

        vacancies.append({
            "id": vac_id,
            "srNo": to_number(cell(ws, row, COL["srNo"])) or None,
            "qty": qty,
            "employeeName": (employee if has_name else None),
            "position": position,
            "company": company,
            "department": department,
            "collarType": cell(ws, row, COL["collar"]),
            "employmentType": cell(ws, row, COL["employment"]),
            "localOverseas": cell(ws, row, COL["localOverseas"]),
            "source": cell(ws, row, COL["source"]),
            "status": status,
            "hiringDate": hiring_iso,
            "remarks": cell(ws, row, REMARKS_COL),
            "costs": {
                "oneTime": one_time,
                "monthlyRecurring": monthly_rec,
                "twoYear": two_year,
                "annual": annual,
                "monthly": monthly,
            },
            "originalData": original,
        })

    data = {
        "metadata": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "source": input_path,
            "version": "1.0",
            "currency": "AED",
            "recordCount": len(vacancies),
            "costRules": {
                "oneTime": "Columns M-O — charged once in the hiring month.",
                "monthlyRecurring": "Columns P-R + Y — charged every month from the hiring month.",
                "twoYear": "Columns S & Z — charged in the hiring month, then every 24 months.",
                "annual": "Columns T-X — charged in the hiring month, then every 12 months.",
            },
        },
        "columns": columns_meta,
        "vacancies": vacancies,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)

    # ---- Report ----------------------------------------------------------
    print(f"OK  Converted {len(vacancies)} vacancies -> {output_path}")
    companies = sorted({v['company'] for v in vacancies if v['company']})
    departments = sorted({v['department'] for v in vacancies if v['department']})
    print(f"    Companies   ({len(companies)}): {', '.join(map(str, companies))}")
    print(f"    Departments ({len(departments)}): {', '.join(map(str, departments))}")
    if warnings:
        print(f"\n{len(warnings)} data-quality warning(s):")
        for w in warnings[:40]:
            print(f"    - {w}")
        if len(warnings) > 40:
            print(f"    ... and {len(warnings) - 40} more")
    else:
        print("    No data-quality issues detected.")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit("ERROR: provide the input Excel file.  e.g.  python convert.py hiring.xlsx")
    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "data.json"
    try:
        convert(input_path, output_path)
    except FileNotFoundError:
        sys.exit(f"ERROR: file not found: {input_path}")
    except Exception as exc:
        sys.exit(f"ERROR during conversion: {exc}")


if __name__ == "__main__":
    main()
